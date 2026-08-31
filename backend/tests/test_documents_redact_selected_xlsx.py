import io
import zipfile

import docx
import openpyxl
import pymupdf
from fastapi.testclient import TestClient

from backend.app.main import app

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats"
    "-officedocument.spreadsheetml.sheet"
)


def upload_xlsx(client, workbook):
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/documents",
        files={
            "file": (
                "sample.xlsx",
                buffer,
                XLSX_CONTENT_TYPE,
            )
        },
    )

    assert response.status_code == 200

    return response.json()["id"]


def _no_ner(monkeypatch):
    monkeypatch.setattr(
        "backend.app.services.detection_engine.detect_named_entities",
        lambda text: [],
    )


def fake_location_ner(text):
    start = text.index("İstanbul")
    end = start + len("İstanbul")

    return [
        {
            "type": "LOCATION",
            "value": "İstanbul",
            "start": start,
            "end": end,
            "confidence": 0.99,
            "source": "ner_model",
        }
    ]


def test_redact_selected_xlsx_auto_redacts_email(monkeypatch):
    _no_ner(monkeypatch)

    client = TestClient(app)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "E-mail: test@example.com ile iletişime geçin."

    document_id = upload_xlsx(client, workbook)

    response = client.post(
        f"/api/documents/{document_id}/redact-selected",
        json={"selected_finding_ids": []},
    )

    assert response.status_code == 200

    redacted = openpyxl.load_workbook(
        io.BytesIO(response.content)
    )
    redacted_value = redacted.active["A1"].value

    assert "test@example.com" not in redacted_value
    assert "ile iletişime geçin." in redacted_value
    assert "t**t@e*****e.com" in redacted_value


def test_redact_selected_xlsx_keeps_unselected_review_finding(
    monkeypatch,
):
    monkeypatch.setattr(
        "backend.app.services.detection_engine.detect_named_entities",
        fake_location_ner,
    )

    client = TestClient(app)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = (
        "Merkezimiz İstanbul şehrindedir. "
        "Email: test@example.com"
    )

    document_id = upload_xlsx(client, workbook)

    response = client.post(
        f"/api/documents/{document_id}/redact-selected",
        json={"selected_finding_ids": []},
    )

    assert response.status_code == 200

    redacted = openpyxl.load_workbook(
        io.BytesIO(response.content)
    )
    redacted_value = redacted.active["A1"].value

    assert "İstanbul" in redacted_value
    assert "test@example.com" not in redacted_value


def test_redact_selected_xlsx_redacts_selected_review_finding(
    monkeypatch,
):
    monkeypatch.setattr(
        "backend.app.services.detection_engine.detect_named_entities",
        fake_location_ner,
    )

    client = TestClient(app)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = (
        "Merkezimiz İstanbul şehrindedir. "
        "Email: test@example.com"
    )

    document_id = upload_xlsx(client, workbook)

    analyze_response = client.get(
        f"/api/documents/{document_id}/analyze"
    )

    assert analyze_response.status_code == 200

    findings = analyze_response.json()["findings"]

    location_finding = next(
        finding
        for finding in findings
        if finding["type"] == "LOCATION"
    )

    response = client.post(
        f"/api/documents/{document_id}/redact-selected",
        json={
            "selected_finding_ids": [
                location_finding["finding_id"]
            ]
        },
    )

    assert response.status_code == 200

    redacted = openpyxl.load_workbook(
        io.BytesIO(response.content)
    )
    redacted_value = redacted.active["A1"].value

    assert "İstanbul" not in redacted_value
    assert "test@example.com" not in redacted_value


def test_redact_selected_xlsx_redacts_finding_on_second_sheet(
    monkeypatch,
):
    _no_ner(monkeypatch)

    client = TestClient(app)

    workbook = openpyxl.Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Genel"
    first_sheet["A1"] = "Hassas veri içermez"

    second_sheet = workbook.create_sheet("İletişim")
    second_sheet["A1"] = "E-mail: test@example.com"

    document_id = upload_xlsx(client, workbook)

    response = client.post(
        f"/api/documents/{document_id}/redact-selected",
        json={"selected_finding_ids": []},
    )

    assert response.status_code == 200

    redacted = openpyxl.load_workbook(
        io.BytesIO(response.content)
    )

    assert (
        redacted["Genel"]["A1"].value
        == "Hassas veri içermez"
    )
    assert (
        "test@example.com"
        not in redacted["İletişim"]["A1"].value
    )
    assert "t**t@e*****e.com" in redacted["İletişim"]["A1"].value


def test_redact_selected_xlsx_response_content_type(monkeypatch):
    _no_ner(monkeypatch)

    client = TestClient(app)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "E-mail: test@example.com"

    document_id = upload_xlsx(client, workbook)

    response = client.post(
        f"/api/documents/{document_id}/redact-selected",
        json={"selected_finding_ids": []},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX_CONTENT_TYPE
    assert response.headers["content-disposition"].endswith(
        '.xlsx"'
    )


def test_redact_selected_xlsx_reopened_workbook_has_no_sensitive_text(
    monkeypatch,
):
    _no_ner(monkeypatch)

    client = TestClient(app)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "İletişim: test@example.com numaralı kayıt."

    document_id = upload_xlsx(client, workbook)

    response = client.post(
        f"/api/documents/{document_id}/redact-selected",
        json={"selected_finding_ids": []},
    )

    assert response.status_code == 200

    workbook_reopened = openpyxl.load_workbook(
        io.BytesIO(response.content)
    )
    redacted_value = workbook_reopened.active["A1"].value

    assert "test@example.com" not in redacted_value
    assert "İletişim:" in redacted_value
    assert "numaralı kayıt." in redacted_value


def test_redact_selected_xlsx_removes_value_from_raw_zip_xml(
    monkeypatch,
):
    _no_ner(monkeypatch)

    client = TestClient(app)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "ayfer.aycan@example.com adresine yazın"

    document_id = upload_xlsx(client, workbook)

    response = client.post(
        f"/api/documents/{document_id}/redact-selected",
        json={"selected_finding_ids": []},
    )

    assert response.status_code == 200

    with zipfile.ZipFile(
        io.BytesIO(response.content)
    ) as archive:
        xml_entries = [
            name
            for name in archive.namelist()
            if name.endswith(".xml")
        ]

        assert "xl/worksheets/sheet1.xml" in xml_entries

        for name in xml_entries:
            content = archive.read(name).decode(
                "utf-8", errors="replace"
            )

            assert "ayfer.aycan@example.com" not in content

        worksheet_xml = archive.read(
            "xl/worksheets/sheet1.xml"
        ).decode("utf-8")

    assert "a*********n@e*****e.com" in worksheet_xml
    assert "adresine yaz" in worksheet_xml


def test_redact_selected_xlsx_formula_cell_fail_safe(monkeypatch):
    def fake_person_ner(text):
        value = "Ayşe Yılmaz"

        if value not in text:
            return []

        start = text.index(value)

        return [
            {
                "type": "PERSON",
                "value": value,
                "start": start,
                "end": start + len(value),
                "confidence": 0.99,
                "source": "ner_model",
            }
        ]

    monkeypatch.setattr(
        "backend.app.services.detection_engine.detect_named_entities",
        fake_person_ner,
    )

    client = TestClient(app)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Ayşe"
    sheet["B1"] = "Yılmaz"
    sheet["C1"] = '=A1&" "&B1'

    upload_buffer = io.BytesIO()
    workbook.save(upload_buffer)
    upload_buffer.seek(0)

    # openpyxl never computes formulas, so patch a cached result into the
    # sheet XML directly (as a real spreadsheet application would) to
    # exercise the fail-safe formula-cell path deterministically.
    with zipfile.ZipFile(upload_buffer) as archive:
        entries = {
            name: archive.read(name)
            for name in archive.namelist()
        }

    sheet_xml = entries["xl/worksheets/sheet1.xml"].decode(
        "utf-8"
    )
    empty_cell = '<c r="C1"><f>A1&amp;" "&amp;B1</f><v></v></c>'
    cached_cell = (
        '<c r="C1" t="str"><f>A1&amp;" "&amp;B1</f>'
        "<v>Ayşe Yılmaz</v></c>"
    )
    assert empty_cell in sheet_xml
    entries["xl/worksheets/sheet1.xml"] = sheet_xml.replace(
        empty_cell, cached_cell
    ).encode("utf-8")

    patched_buffer = io.BytesIO()
    with zipfile.ZipFile(
        patched_buffer, "w", zipfile.ZIP_DEFLATED
    ) as archive:
        for name, data in entries.items():
            archive.writestr(name, data)
    patched_buffer.seek(0)

    upload_response = client.post(
        "/api/documents",
        files={
            "file": (
                "formula.xlsx",
                patched_buffer,
                XLSX_CONTENT_TYPE,
            )
        },
    )

    assert upload_response.status_code == 200

    document_id = upload_response.json()["id"]

    analyze_response = client.get(
        f"/api/documents/{document_id}/analyze"
    )

    assert analyze_response.status_code == 200

    findings = analyze_response.json()["findings"]

    assert any(
        finding["value"] == "Ayşe Yılmaz"
        for finding in findings
    )

    response = client.post(
        f"/api/documents/{document_id}/redact-selected",
        json={"selected_finding_ids": []},
    )

    assert response.status_code == 200

    redacted_formula_mode = openpyxl.load_workbook(
        io.BytesIO(response.content),
        data_only=False,
    )
    redacted_cell = redacted_formula_mode.active["C1"]

    assert redacted_cell.data_type != "f"
    assert "=" not in redacted_cell.value
    assert redacted_cell.value == "A**e Y****z"

    with zipfile.ZipFile(
        io.BytesIO(response.content)
    ) as archive:
        raw_xml = archive.read(
            "xl/worksheets/sheet1.xml"
        ).decode("utf-8")

    assert "Ayşe Yılmaz" not in raw_xml
    assert "A**e Y****z" in raw_xml
    assert "A1&amp;" not in raw_xml


def test_redact_selected_pdf_and_docx_are_unaffected_by_xlsx_support(
    monkeypatch,
):
    """Regression guard: PDF/DOCX dispatch still works unchanged."""
    _no_ner(monkeypatch)

    client = TestClient(app)

    pdf_document = pymupdf.open()
    page = pdf_document.new_page()
    page.insert_text((72, 72), "E-mail: test@example.com")

    pdf_buffer = io.BytesIO()
    pdf_document.save(pdf_buffer)
    pdf_document.close()
    pdf_buffer.seek(0)

    pdf_upload = client.post(
        "/api/documents",
        files={
            "file": ("sample.pdf", pdf_buffer, "application/pdf")
        },
    )

    pdf_document_id = pdf_upload.json()["id"]

    pdf_response = client.post(
        f"/api/documents/{pdf_document_id}/redact-selected",
        json={"selected_finding_ids": []},
    )

    assert pdf_response.status_code == 200
    assert pdf_response.headers["content-type"] == "application/pdf"

    docx_document = docx.Document()
    docx_document.add_paragraph("E-mail: test@example.com")

    docx_buffer = io.BytesIO()
    docx_document.save(docx_buffer)
    docx_buffer.seek(0)

    docx_upload = client.post(
        "/api/documents",
        files={
            "file": (
                "sample.docx",
                docx_buffer,
                (
                    "application/vnd.openxmlformats"
                    "-officedocument.wordprocessingml"
                    ".document"
                ),
            )
        },
    )

    docx_document_id = docx_upload.json()["id"]

    docx_response = client.post(
        f"/api/documents/{docx_document_id}/redact-selected",
        json={"selected_finding_ids": []},
    )

    assert docx_response.status_code == 200
    assert docx_response.headers["content-type"] == (
        "application/vnd.openxmlformats"
        "-officedocument.wordprocessingml.document"
    )
