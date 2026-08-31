import zipfile

import openpyxl
from openpyxl.styles import Font, PatternFill

from backend.app.services.xlsx_parser import extract_text_from_xlsx
from backend.app.services.xlsx_redactor import create_redacted_xlsx


def _finding_for(
    page_text: str,
    value: str,
    finding_type: str,
    page_number: int = 1,
) -> dict:
    start = page_text.index(value)

    return {
        "type": finding_type,
        "value": value,
        "page_number": page_number,
        "start": start,
        "end": start + len(value),
    }


def test_create_redacted_xlsx_single_cell_partial_redaction(tmp_path):
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "redacted.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Email: ayse@example.com kayıtlı"
    workbook.save(source_path)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    value = "ayse@example.com"

    create_redacted_xlsx(
        source_path,
        output_path,
        [_finding_for(text, value, "EMAIL")],
    )

    redacted = openpyxl.load_workbook(output_path)
    cell_value = redacted.active["A1"].value

    assert cell_value == (
        "Email: a**e@e*****e.com kayıtlı"
    )
    assert value not in cell_value


def test_create_redacted_xlsx_preserves_neighboring_text(tmp_path):
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "redacted.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = (
        "Ad Soyad: Ayfer Aycan, Email: test@example.com, "
        "Şehir: İstanbul"
    )
    workbook.save(source_path)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    value = "test@example.com"

    create_redacted_xlsx(
        source_path,
        output_path,
        [_finding_for(text, value, "EMAIL")],
    )

    redacted_value = openpyxl.load_workbook(output_path).active[
        "A1"
    ].value

    assert "Ad Soyad: Ayfer Aycan" in redacted_value
    assert "Şehir: İstanbul" in redacted_value
    assert value not in redacted_value


def test_create_redacted_xlsx_masks_person_in_normal_cell(tmp_path):
    source_path = tmp_path / "person-source.xlsx"
    output_path = tmp_path / "person-redacted.xlsx"

    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "Kişi: Ayşe Yılmaz kayıtlı"
    workbook.save(source_path)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    value = "Ayşe Yılmaz"
    create_redacted_xlsx(
        source_path,
        output_path,
        [_finding_for(text, value, "PERSON")],
    )

    assert (
        openpyxl.load_workbook(output_path).active["A1"].value
        == "Kişi: A**e Y****z kayıtlı"
    )


def test_create_redacted_xlsx_multiple_findings_same_cell(tmp_path):
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "redacted.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Email: test@example.com Phone: 05321234567"
    workbook.save(source_path)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    email = "test@example.com"
    phone = "05321234567"

    findings = [
        _finding_for(text, email, "EMAIL"),
        _finding_for(text, phone, "PHONE"),
    ]

    create_redacted_xlsx(source_path, output_path, findings)

    redacted_value = openpyxl.load_workbook(output_path).active[
        "A1"
    ].value

    assert email not in redacted_value
    assert phone not in redacted_value
    assert "Email: " in redacted_value
    assert " Phone: " in redacted_value
    assert redacted_value == (
        "Email: "
        + "t**t@e*****e.com"
        + " Phone: "
        + "0*********7"
    )


def test_create_redacted_xlsx_masks_card_number(tmp_path):
    source_path = tmp_path / "card-source.xlsx"
    output_path = tmp_path / "card-redacted.xlsx"
    value = "4111-1111-1111-1111"

    workbook = openpyxl.Workbook()
    workbook.active["A1"] = f"Kart: {value}"
    workbook.save(source_path)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    create_redacted_xlsx(
        source_path,
        output_path,
        [_finding_for(text, value, "CARD_NUMBER")],
    )

    assert (
        openpyxl.load_workbook(output_path).active["A1"].value
        == "Kart: 4***-****-****-***1"
    )


def test_overlapping_findings_keep_the_more_restricted_mask(tmp_path):
    source_path = tmp_path / "overlap-source.xlsx"
    output_path = tmp_path / "overlap-redacted.xlsx"
    value = "12345678901"

    workbook = openpyxl.Workbook()
    workbook.active["A1"] = value
    workbook.save(source_path)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    create_redacted_xlsx(
        source_path,
        output_path,
        [
            _finding_for(text, value, "TCKN"),
            _finding_for(text, value, "PHONE"),
        ],
    )

    assert (
        openpyxl.load_workbook(output_path).active["A1"].value
        == "***********"
    )


def test_create_redacted_xlsx_multi_cell_range(tmp_path):
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "redacted.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Hello"
    sheet["B1"] = "World"
    workbook.save(source_path)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    assert text == "Hello | World"

    start = text.index("llo")
    end = text.index("Wor") + len("Wor")

    create_redacted_xlsx(
        source_path,
        output_path,
        [
            {
                "type": "PERSON",
                "value": text[start:end],
                "page_number": 1,
                "start": start,
                "end": end,
            }
        ],
    )

    redacted_sheet = openpyxl.load_workbook(output_path).active

    assert redacted_sheet["A1"].value == "Hel*o"
    assert redacted_sheet["B1"].value == "W*rld"


def test_create_redacted_xlsx_second_sheet(tmp_path):
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "redacted.xlsx"

    workbook = openpyxl.Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Genel"
    first_sheet["A1"] = "test@example.com"

    second_sheet = workbook.create_sheet("İletişim")
    second_sheet["A1"] = "test@example.com"

    workbook.save(source_path)

    pages = extract_text_from_xlsx(source_path)["pages"]
    second_page_text = pages[1]["text"]
    value = "test@example.com"

    create_redacted_xlsx(
        source_path,
        output_path,
        [_finding_for(second_page_text, value, "EMAIL", page_number=2)],
    )

    redacted = openpyxl.load_workbook(output_path)

    assert redacted["Genel"]["A1"].value == "test@example.com"
    assert redacted["İletişim"]["A1"].value == "t**t@e*****e.com"


def test_create_redacted_xlsx_merged_cell_no_crash(tmp_path):
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "redacted.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Başlık"
    sheet.merge_cells("A1:C1")
    sheet["A2"] = "Ad"
    sheet["B2"] = "test@example.com"
    workbook.save(source_path)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    value = "test@example.com"

    create_redacted_xlsx(
        source_path,
        output_path,
        [_finding_for(text, value, "EMAIL")],
    )

    redacted = openpyxl.load_workbook(output_path)
    redacted_sheet = redacted.active

    assert redacted_sheet["A1"].value == "Başlık"
    assert redacted_sheet["B2"].value == "t**t@e*****e.com"
    # Merged-away placeholders must stay untouched (None), not crash.
    assert redacted_sheet["B1"].value is None
    assert redacted_sheet["C1"].value is None


def test_create_redacted_xlsx_numeric_value_redaction(tmp_path):
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "redacted.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = 12345678901
    workbook.save(source_path)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    assert text == "12345678901"

    create_redacted_xlsx(
        source_path,
        output_path,
        [
            {
                "type": "TCKN",
                "value": text,
                "page_number": 1,
                "start": 0,
                "end": len(text),
            }
        ],
    )

    redacted_value = openpyxl.load_workbook(output_path).active[
        "A1"
    ].value

    assert redacted_value == "***********"
    assert isinstance(redacted_value, str)


def test_create_redacted_xlsx_formula_cell_whole_cell_redaction(
    tmp_path,
):
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "redacted.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Ayşe"
    sheet["B1"] = "Yılmaz"
    sheet["C1"] = '=A1&" "&B1'
    workbook.save(source_path)

    # openpyxl never computes formulas, so patch in a cached result the
    # way a real spreadsheet application would, directly in the sheet
    # XML, to exercise the fail-safe formula-cell path deterministically.
    with zipfile.ZipFile(source_path) as archive:
        entries = {
            item.filename: archive.read(item.filename)
            for item in archive.infolist()
        }

    sheet_xml = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
    empty_cell = (
        '<c r="C1"><f>A1&amp;" "&amp;B1</f><v></v></c>'
    )
    cached_cell = (
        '<c r="C1" t="str"><f>A1&amp;" "&amp;B1</f>'
        "<v>Ayşe Yılmaz</v></c>"
    )
    assert empty_cell in sheet_xml
    sheet_xml = sheet_xml.replace(empty_cell, cached_cell)
    entries["xl/worksheets/sheet1.xml"] = sheet_xml.encode("utf-8")

    with zipfile.ZipFile(
        source_path, "w", zipfile.ZIP_DEFLATED
    ) as archive:
        for name, data in entries.items():
            archive.writestr(name, data)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    assert text == "Ayşe | Yılmaz | Ayşe Yılmaz"

    cached_value = "Ayşe Yılmaz"
    finding = _finding_for(text, cached_value, "PERSON")

    create_redacted_xlsx(source_path, output_path, [finding])

    redacted_formula_mode = openpyxl.load_workbook(
        output_path, data_only=False
    )
    redacted_cell = redacted_formula_mode.active["C1"]

    assert redacted_cell.data_type != "f"
    assert redacted_cell.value == "A**e Y****z"
    assert "=" not in redacted_cell.value

    redacted_data_mode = openpyxl.load_workbook(
        output_path, data_only=True
    )

    assert (
        redacted_data_mode.active["C1"].value
        == "A**e Y****z"
    )

    with zipfile.ZipFile(output_path) as archive:
        raw_xml = archive.read(
            "xl/worksheets/sheet1.xml"
        ).decode("utf-8")

    assert cached_value not in raw_xml
    assert "A**e Y****z" in raw_xml
    assert "A1&amp;" not in raw_xml


def test_create_redacted_xlsx_reopened_workbook_has_no_sensitive_text(
    tmp_path,
):
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "redacted.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "TCKN: 98765432109 numaralı kayıt."
    workbook.save(source_path)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    value = "98765432109"

    create_redacted_xlsx(
        source_path,
        output_path,
        [_finding_for(text, value, "TCKN")],
    )

    redacted_text = extract_text_from_xlsx(output_path)["pages"][
        0
    ]["text"]

    assert value not in redacted_text
    assert "TCKN:" in redacted_text
    assert "numaralı kayıt." in redacted_text


def test_create_redacted_xlsx_removes_value_from_raw_zip_xml(
    tmp_path,
):
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "redacted.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "ayfer.aycan@example.com adresine yazın"
    workbook.save(source_path)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    value = "ayfer.aycan@example.com"

    create_redacted_xlsx(
        source_path,
        output_path,
        [_finding_for(text, value, "EMAIL")],
    )

    with zipfile.ZipFile(output_path) as archive:
        xml_entries = [
            name
            for name in archive.namelist()
            if name.endswith(".xml")
        ]

        assert "xl/worksheets/sheet1.xml" in xml_entries

        xml_contents = []

        for name in xml_entries:
            content = archive.read(name).decode(
                "utf-8", errors="replace"
            )
            xml_contents.append(content)

            assert value not in content

    combined_xml = "\n".join(xml_contents)
    assert "a*********n@e*****e.com" in combined_xml
    assert "adresine yaz" in combined_xml


def test_create_redacted_xlsx_preserves_cell_style(tmp_path):
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "redacted.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "test@example.com"
    sheet["A1"].font = Font(bold=True, color="FF0000")
    sheet["A1"].fill = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type="solid",
    )
    workbook.save(source_path)

    text = extract_text_from_xlsx(source_path)["pages"][0]["text"]
    value = "test@example.com"

    create_redacted_xlsx(
        source_path,
        output_path,
        [_finding_for(text, value, "EMAIL")],
    )

    redacted_cell = openpyxl.load_workbook(output_path).active[
        "A1"
    ]

    assert redacted_cell.value == "t**t@e*****e.com"
    assert redacted_cell.font.bold is True
    assert redacted_cell.font.color.rgb == "00FF0000"
    assert redacted_cell.fill.start_color.rgb == "00FFFF00"
