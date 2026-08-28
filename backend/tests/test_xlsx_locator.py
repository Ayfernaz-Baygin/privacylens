import zipfile

import openpyxl

from backend.app.services.xlsx_locator import (
    build_xlsx_cell_index,
    locate_text_in_xlsx,
)
from backend.app.services.xlsx_parser import extract_text_from_xlsx


def test_locate_text_in_xlsx_single_cell(tmp_path):
    xlsx_path = tmp_path / "single-cell.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "E-mail: test@example.com"
    workbook.save(xlsx_path)

    text = extract_text_from_xlsx(xlsx_path)["pages"][0]["text"]
    value = "test@example.com"
    start = text.index(value)
    end = start + len(value)

    matches = locate_text_in_xlsx(
        xlsx_path,
        page_number=1,
        start=start,
        end=end,
    )

    assert len(matches) == 1

    match = matches[0]

    assert match["coordinate"] == "A1"
    assert match["sheet_name"] == "Sheet"
    assert match["matched_text"] == value


def test_locate_text_in_xlsx_match_offsets_are_cell_local(tmp_path):
    xlsx_path = tmp_path / "local-offsets.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Prefix"
    sheet["B1"] = "test@example.com"
    workbook.save(xlsx_path)

    text = extract_text_from_xlsx(xlsx_path)["pages"][0]["text"]
    assert text == "Prefix | test@example.com"

    value = "example"
    start = text.index(value)
    end = start + len(value)

    matches = locate_text_in_xlsx(
        xlsx_path,
        page_number=1,
        start=start,
        end=end,
    )

    assert len(matches) == 1

    match = matches[0]

    assert match["coordinate"] == "B1"
    assert match["match_start"] == 5
    assert match["match_end"] == 12
    assert match["matched_text"] == "example"
    assert match["cell_text"][
        match["match_start"] : match["match_end"]
    ] == "example"


def test_locate_text_in_xlsx_second_sheet(tmp_path):
    xlsx_path = tmp_path / "second-sheet.xlsx"

    workbook = openpyxl.Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Genel"
    first_sheet["A1"] = "Hassas veri yok"

    second_sheet = workbook.create_sheet("İletişim")
    second_sheet["A1"] = "E-mail: test@example.com"

    workbook.save(xlsx_path)

    pages = extract_text_from_xlsx(xlsx_path)["pages"]
    second_page_text = pages[1]["text"]

    value = "test@example.com"
    start = second_page_text.index(value)
    end = start + len(value)

    matches = locate_text_in_xlsx(
        xlsx_path,
        page_number=2,
        start=start,
        end=end,
    )

    assert len(matches) == 1

    match = matches[0]

    assert match["page_number"] == 2
    assert match["sheet_name"] == "İletişim"
    assert match["coordinate"] == "A1"
    assert match["matched_text"] == value

    # A page_number=1 lookup must only ever see sheet "Genel" cells.
    first_page_matches = locate_text_in_xlsx(
        xlsx_path,
        page_number=1,
        start=0,
        end=len("Hassas"),
    )

    assert len(first_page_matches) == 1
    assert first_page_matches[0]["sheet_name"] == "Genel"


def test_locate_text_in_xlsx_offset_parity_with_empty_cells(tmp_path):
    xlsx_path = tmp_path / "empty-cells.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Ad"
    sheet["B1"] = None
    sheet["C1"] = "Soyad"
    # Row 2 is left entirely empty on purpose.
    sheet["A3"] = "test@example.com"
    workbook.save(xlsx_path)

    text = extract_text_from_xlsx(xlsx_path)["pages"][0]["text"]
    assert text == "Ad | Soyad\ntest@example.com"

    value = "test@example.com"
    start = text.index(value)
    end = start + len(value)

    matches = locate_text_in_xlsx(
        xlsx_path,
        page_number=1,
        start=start,
        end=end,
    )

    assert len(matches) == 1
    assert matches[0]["coordinate"] == "A3"
    assert matches[0]["matched_text"] == value


def test_locate_text_in_xlsx_range_spanning_two_cells(tmp_path):
    xlsx_path = tmp_path / "two-cells.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Hello"
    sheet["B1"] = "World"
    workbook.save(xlsx_path)

    text = extract_text_from_xlsx(xlsx_path)["pages"][0]["text"]
    assert text == "Hello | World"

    # Spans from inside "Hello" across the " | " separator into "World".
    start = text.index("llo")
    end = text.index("Wor") + len("Wor")

    matches = locate_text_in_xlsx(
        xlsx_path,
        page_number=1,
        start=start,
        end=end,
    )

    assert len(matches) == 2

    first, second = matches

    assert first["coordinate"] == "A1"
    assert first["matched_text"] == "llo"

    assert second["coordinate"] == "B1"
    assert second["matched_text"] == "Wor"

    assert "|" not in first["matched_text"]
    assert "|" not in second["matched_text"]


def test_locate_text_in_xlsx_does_not_crash_on_merged_cells(tmp_path):
    xlsx_path = tmp_path / "merged.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Başlık"
    sheet.merge_cells("A1:C1")
    sheet["A2"] = "Ad"
    sheet["B2"] = "test@example.com"
    workbook.save(xlsx_path)

    text = extract_text_from_xlsx(xlsx_path)["pages"][0]["text"]
    value = "test@example.com"
    start = text.index(value)
    end = start + len(value)

    matches = locate_text_in_xlsx(
        xlsx_path,
        page_number=1,
        start=start,
        end=end,
    )

    assert len(matches) == 1
    assert matches[0]["coordinate"] == "B2"
    assert matches[0]["matched_text"] == value


def test_build_xlsx_cell_index_text_matches_extract_text_from_xlsx(
    tmp_path,
):
    xlsx_path = tmp_path / "parity.xlsx"

    workbook = openpyxl.Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Müşteriler"
    first_sheet["A1"] = "Ad Soyad"
    first_sheet["B1"] = "Ayşe Yılmaz"
    first_sheet["A3"] = "İkinci satır"

    second_sheet = workbook.create_sheet("Çalışanlar")
    second_sheet["A1"] = "test@example.com"

    workbook.save(xlsx_path)

    parsed_pages = extract_text_from_xlsx(xlsx_path)["pages"]
    index_pages = build_xlsx_cell_index(xlsx_path)["pages"]

    assert len(parsed_pages) == len(index_pages)

    for parsed_page, index_page in zip(
        parsed_pages, index_pages
    ):
        assert parsed_page["text"] == index_page["text"]
        assert (
            parsed_page["sheet_name"]
            == index_page["sheet_name"]
        )
        assert (
            parsed_page["page_number"]
            == index_page["page_number"]
        )


def test_locate_text_in_xlsx_reports_formula_metadata(tmp_path):
    xlsx_path = tmp_path / "formula.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = 5
    sheet["A2"] = 10
    sheet["A3"] = "=A1+A2"
    workbook.save(xlsx_path)

    with zipfile.ZipFile(xlsx_path) as archive:
        sheet_xml = archive.read(
            "xl/worksheets/sheet1.xml"
        ).decode("utf-8")
        entries = {
            item.filename: archive.read(item.filename)
            for item in archive.infolist()
        }

    sheet_xml = sheet_xml.replace(
        '<c r="A3"><f>A1+A2</f><v></v></c>',
        '<c r="A3"><f>A1+A2</f><v>15</v></c>',
    )
    entries["xl/worksheets/sheet1.xml"] = sheet_xml.encode(
        "utf-8"
    )

    with zipfile.ZipFile(
        xlsx_path, "w", zipfile.ZIP_DEFLATED
    ) as archive:
        for name, data in entries.items():
            archive.writestr(name, data)

    text = extract_text_from_xlsx(xlsx_path)["pages"][0]["text"]
    assert "15" in text

    start = text.index("15")
    end = start + len("15")

    matches = locate_text_in_xlsx(
        xlsx_path,
        page_number=1,
        start=start,
        end=end,
    )

    assert len(matches) == 1

    match = matches[0]

    assert match["coordinate"] == "A3"
    assert match["matched_text"] == "15"
    assert match["is_formula"] is True
    assert match["formula"] == "=A1+A2"

    plain_cell_matches = locate_text_in_xlsx(
        xlsx_path,
        page_number=1,
        start=0,
        end=1,
    )

    assert plain_cell_matches[0]["is_formula"] is False
    assert plain_cell_matches[0]["formula"] is None
