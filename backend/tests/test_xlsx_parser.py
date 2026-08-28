import datetime

import openpyxl

from backend.app.services.xlsx_parser import extract_text_from_xlsx


def test_extract_text_from_xlsx_single_sheet(tmp_path):
    xlsx_path = tmp_path / "sample.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Müşteriler"
    sheet["A1"] = "Merhaba dünya"
    workbook.save(xlsx_path)

    result = extract_text_from_xlsx(xlsx_path)

    assert result["page_count"] == 1
    assert len(result["pages"]) == 1

    page = result["pages"][0]

    assert page["page_number"] == 1
    assert page["sheet_name"] == "Müşteriler"
    assert page["text"] == "Merhaba dünya"


def test_extract_text_from_xlsx_preserves_sheet_order(tmp_path):
    xlsx_path = tmp_path / "sample-order.xlsx"

    workbook = openpyxl.Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Müşteriler"
    first_sheet["A1"] = "İlk sayfa"

    second_sheet = workbook.create_sheet("Çalışanlar")
    second_sheet["A1"] = "İkinci sayfa"

    third_sheet = workbook.create_sheet("Raporlar")
    third_sheet["A1"] = "Üçüncü sayfa"

    workbook.save(xlsx_path)

    result = extract_text_from_xlsx(xlsx_path)

    assert result["page_count"] == 3

    page_numbers = [
        page["page_number"] for page in result["pages"]
    ]
    sheet_names = [
        page["sheet_name"] for page in result["pages"]
    ]
    texts = [page["text"] for page in result["pages"]]

    assert page_numbers == [1, 2, 3]
    assert sheet_names == [
        "Müşteriler",
        "Çalışanlar",
        "Raporlar",
    ]
    assert texts == [
        "İlk sayfa",
        "İkinci sayfa",
        "Üçüncü sayfa",
    ]


def test_extract_text_from_xlsx_skips_empty_cells_and_rows(tmp_path):
    xlsx_path = tmp_path / "sample-empty.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Ad"
    sheet["B1"] = None
    sheet["C1"] = "Soyad"
    # Row 2 is left entirely empty on purpose.
    sheet["A3"] = "Ayfer"
    workbook.save(xlsx_path)

    result = extract_text_from_xlsx(xlsx_path)

    page = result["pages"][0]
    lines = page["text"].split("\n")

    assert lines == ["Ad | Soyad", "Ayfer"]


def test_extract_text_from_xlsx_joins_row_cells_with_pipe(tmp_path):
    xlsx_path = tmp_path / "sample-row.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Ad Soyad"
    sheet["B1"] = "Ayşe Yılmaz"
    workbook.save(xlsx_path)

    result = extract_text_from_xlsx(xlsx_path)

    assert result["pages"][0]["text"] == (
        "Ad Soyad | Ayşe Yılmaz"
    )


def test_extract_text_from_xlsx_converts_numbers_dates_and_booleans(
    tmp_path,
):
    xlsx_path = tmp_path / "sample-types.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = 42
    sheet["B1"] = 3.14
    sheet["C1"] = True
    sheet["D1"] = datetime.date(2026, 8, 28)
    workbook.save(xlsx_path)

    result = extract_text_from_xlsx(xlsx_path)

    text = result["pages"][0]["text"]
    cells = text.split(" | ")

    assert cells[0] == "42"
    assert cells[1] == "3.14"
    assert cells[2] == "True"
    assert "2026-08-28" in cells[3]


def test_extract_text_from_xlsx_uses_cached_formula_value(tmp_path):
    xlsx_path = tmp_path / "sample-formula.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = 5
    sheet["A2"] = 10
    sheet["A3"] = "=A1+A2"
    workbook.save(xlsx_path)

    result = extract_text_from_xlsx(xlsx_path)

    text = result["pages"][0]["text"]

    assert "=A1+A2" not in text
    assert "SUM" not in text.upper()


def test_extract_text_from_xlsx_does_not_crash_on_merged_cells(
    tmp_path,
):
    xlsx_path = tmp_path / "sample-merged.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Başlık"
    sheet.merge_cells("A1:C1")
    sheet["A2"] = "Ad"
    sheet["B2"] = "Soyad"
    workbook.save(xlsx_path)

    result = extract_text_from_xlsx(xlsx_path)

    page = result["pages"][0]
    lines = page["text"].split("\n")

    assert lines == ["Başlık", "Ad | Soyad"]
