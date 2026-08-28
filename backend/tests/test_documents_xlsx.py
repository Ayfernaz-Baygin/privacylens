import io

import openpyxl
from fastapi.testclient import TestClient

from backend.app.main import app

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats"
    "-officedocument.spreadsheetml.sheet"
)


def upload_xlsx(client, workbook):
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/documents",
        files={
            "file": (
                "sample.xlsx",
                buffer,
                XLSX_CONTENT_TYPE,
            )
        },
    )

    assert response.status_code == 200

    return response.json()["id"]


def _no_ner(monkeypatch):
    monkeypatch.setattr(
        "backend.app.services.detection_engine.detect_named_entities",
        lambda text: [],
    )


def test_text_extraction_single_sheet(monkeypatch):
    client = TestClient(app)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Müşteriler"
    sheet["A1"] = "Ad Soyad"
    sheet["B1"] = "Ayşe Yılmaz"

    document_id = upload_xlsx(client, workbook)

    response = client.get(
        f"/api/documents/{document_id}/text"
    )

    assert response.status_code == 200

    body = response.json()

    assert body["status"] == "parsed"
    assert body["page_count"] == 1
    assert len(body["pages"]) == 1

    page = body["pages"][0]

    assert page["page_number"] == 1
    assert page["sheet_name"] == "Müşteriler"
    assert page["text"] == "Ad Soyad | Ayşe Yılmaz"


def test_text_extraction_multi_sheet_order(monkeypatch):
    client = TestClient(app)

    workbook = openpyxl.Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Müşteriler"
    first_sheet["A1"] = "İlk sayfa"

    second_sheet = workbook.create_sheet("Çalışanlar")
    second_sheet["A1"] = "İkinci sayfa"

    document_id = upload_xlsx(client, workbook)

    response = client.get(
        f"/api/documents/{document_id}/text"
    )

    assert response.status_code == 200

    body = response.json()

    assert body["page_count"] == 2

    page_numbers = [
        page["page_number"] for page in body["pages"]
    ]
    sheet_names = [
        page["sheet_name"] for page in body["pages"]
    ]

    assert page_numbers == [1, 2]
    assert sheet_names == ["Müşteriler", "Çalışanlar"]
    assert body["pages"][0]["text"] == "İlk sayfa"
    assert body["pages"][1]["text"] == "İkinci sayfa"


def test_analyze_xlsx_detects_email(monkeypatch):
    _no_ner(monkeypatch)

    client = TestClient(app)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "E-mail: test@example.com"

    document_id = upload_xlsx(client, workbook)

    response = client.get(
        f"/api/documents/{document_id}/analyze"
    )

    assert response.status_code == 200

    body = response.json()

    assert body["status"] == "analyzed"
    assert body["page_count"] == 1
    assert body["finding_count"] == 1

    finding = body["findings"][0]

    assert finding["type"] == "EMAIL"
    assert finding["value"] == "test@example.com"
    assert finding["bounding_boxes"] == []
    assert finding["page_number"] == 1
    assert "finding_id" in finding
    assert "confidence_level" in finding
    assert "privacy_status" in finding
    assert "redaction_action" in finding


def test_analyze_xlsx_finding_page_number_matches_sheet_order(
    monkeypatch,
):
    _no_ner(monkeypatch)

    client = TestClient(app)

    workbook = openpyxl.Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Genel"
    first_sheet["A1"] = "Hassas veri içermez"

    second_sheet = workbook.create_sheet("İletişim")
    second_sheet["A1"] = "E-mail: test@example.com"

    document_id = upload_xlsx(client, workbook)

    response = client.get(
        f"/api/documents/{document_id}/analyze"
    )

    assert response.status_code == 200

    body = response.json()

    assert body["page_count"] == 2
    assert body["finding_count"] == 1

    finding = body["findings"][0]

    assert finding["type"] == "EMAIL"
    assert finding["page_number"] == 2
    assert finding["finding_id"].startswith("2:")


def test_analyze_xlsx_without_findings_returns_empty_list(monkeypatch):
    _no_ner(monkeypatch)

    client = TestClient(app)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Herhangi bir hassas veri içermez."

    document_id = upload_xlsx(client, workbook)

    response = client.get(
        f"/api/documents/{document_id}/analyze"
    )

    assert response.status_code == 200

    body = response.json()

    assert body["finding_count"] == 0
    assert body["findings"] == []


def test_redact_selected_rejects_xlsx_without_touching_pdf_flow(
    monkeypatch,
):
    _no_ner(monkeypatch)

    client = TestClient(app)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "E-mail: test@example.com"

    document_id = upload_xlsx(client, workbook)

    response = client.post(
        f"/api/documents/{document_id}/redact-selected",
        json={"selected_finding_ids": []},
    )

    assert response.status_code == 415
    assert "XLSX" in response.json()["detail"]
