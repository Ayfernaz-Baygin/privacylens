from pathlib import Path

import openpyxl

from backend.app.services.xlsx_parser import iter_included_row_cells

LINE_SEPARATOR_LENGTH = len("\n")
CELL_SEPARATOR_LENGTH = len(" | ")


def _row_cell_spans(row) -> tuple[str, list[dict]]:
    """Row text and cell spans relative to that row's own text.

    Mirrors xlsx_parser's row line construction exactly: only included
    (non-empty) cells contribute, joined with " | ". The separator
    characters themselves never fall inside a cell's [start, end) span,
    so they can never be mistaken for cell content by the locator.
    """
    spans = []
    text_parts = []
    offset = 0

    for position, (cell, text) in enumerate(
        iter_included_row_cells(row)
    ):
        if position > 0:
            offset += CELL_SEPARATOR_LENGTH

        spans.append(
            {
                "cell": cell,
                "start": offset,
                "end": offset + len(text),
                "text": text,
            }
        )

        text_parts.append(text)
        offset += len(text)

    return " | ".join(text_parts), spans


def _load_formula_flags(file_path: Path) -> dict:
    """sheet title -> {coordinate: (is_formula, formula_text)}.

    Requires a second, formula-mode read of the workbook (data_only=
    False): the data_only=True pass used for text/offsets only ever sees
    a formula cell's last cached result, never the formula itself, so
    there is no way to recover "is this a formula" from that pass alone.
    """
    workbook = openpyxl.load_workbook(
        file_path,
        data_only=False,
    )

    try:
        flags_by_sheet = {}

        for sheet in workbook.worksheets:
            sheet_flags = {}

            for row in sheet.iter_rows():
                for cell in row:
                    is_formula = cell.data_type == "f"

                    sheet_flags[cell.coordinate] = (
                        is_formula,
                        cell.value if is_formula else None,
                    )

            flags_by_sheet[sheet.title] = sheet_flags

        return flags_by_sheet

    finally:
        workbook.close()


def _build_page(
    sheet,
    page_number: int,
    formula_flags: dict,
) -> dict:
    lines = []

    for row in sheet.iter_rows():
        row_text, spans = _row_cell_spans(row)

        if row_text:
            lines.append((row_text, spans))

    text_parts = []
    cells = []
    offset = 0

    for line_index, (row_text, spans) in enumerate(lines):
        if line_index > 0:
            offset += LINE_SEPARATOR_LENGTH

        for span in spans:
            cell = span["cell"]

            is_formula, formula_text = formula_flags.get(
                cell.coordinate,
                (False, None),
            )

            cells.append(
                {
                    "page_number": page_number,
                    "sheet_name": sheet.title,
                    "row": cell.row,
                    "column": cell.column,
                    "coordinate": cell.coordinate,
                    "start": offset + span["start"],
                    "end": offset + span["end"],
                    "cell_text": span["text"],
                    "is_formula": is_formula,
                    "formula": formula_text,
                }
            )

        text_parts.append(row_text)
        offset += len(row_text)

    return {
        "page_number": page_number,
        "sheet_name": sheet.title,
        "text": "\n".join(text_parts),
        "cells": cells,
    }


def build_xlsx_cell_index(file_path: Path) -> dict:
    """Per-sheet text plus every included cell's offset within it.

    Each page's "text" is built with the exact same rules as
    extract_text_from_xlsx() (same helper, same join/skip logic), so the
    start/end offsets detect_sensitive_data() produces against that page
    text can be looked up here directly. Cached formula results are read
    the same way the parser reads them (data_only=True); formula
    metadata comes from a separate, second read of the workbook, see
    _load_formula_flags().
    """
    formula_flags = _load_formula_flags(file_path)

    workbook = openpyxl.load_workbook(
        file_path,
        data_only=True,
    )

    try:
        pages = [
            _build_page(
                sheet,
                page_number,
                formula_flags.get(sheet.title, {}),
            )
            for page_number, sheet in enumerate(
                workbook.worksheets,
                start=1,
            )
        ]

    finally:
        workbook.close()

    return {"pages": pages}


def locate_entity_cells(
    page: dict,
    start: int,
    end: int,
) -> list[dict]:
    """Cells overlapping [start, end) within a single page's text.

    A single-cell entity yields one record; an entity whose range spans
    more than one cell (e.g. it was matched across a " | " boundary)
    yields one record per cell it touches, in document order.
    """
    matches = []

    for cell in page["cells"]:
        overlap_start = max(cell["start"], start)
        overlap_end = min(cell["end"], end)

        if overlap_start >= overlap_end:
            continue

        local_start = overlap_start - cell["start"]
        local_end = overlap_end - cell["start"]

        matches.append(
            {
                **{
                    key: value
                    for key, value in cell.items()
                    if key not in ("start", "end", "cell_text")
                },
                "cell_start": cell["start"],
                "cell_end": cell["end"],
                "cell_text": cell["cell_text"],
                "match_start": local_start,
                "match_end": local_end,
                "matched_text": cell["cell_text"][
                    local_start:local_end
                ],
            }
        )

    return matches


def locate_text_in_xlsx(
    file_path: Path,
    page_number: int,
    start: int,
    end: int,
) -> list[dict]:
    index = build_xlsx_cell_index(file_path)

    page = next(
        (
            page
            for page in index["pages"]
            if page["page_number"] == page_number
        ),
        None,
    )

    if page is None:
        return []

    return locate_entity_cells(page, start, end)
